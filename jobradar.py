#!/usr/bin/env python3
"""
JobRadar - busca vagas (LinkedIn, Indeed, Google Jobs) e gera mensagem
personalizada por vaga usando um modelo local via Ollama (opcional).

Autor: Lucas Zafalon
GitHub: https://github.com/LucasZafalon
LinkedIn: linkedin.com/in/lucaszafalon

Este script NÃO envia candidaturas nem preenche formulários em nenhuma
plataforma. Ele apenas busca vagas publicamente e, se você quiser, sugere
um texto de mensagem para você revisar e enviar manualmente.

Pré-requisitos:
  pip install python-jobspy openpyxl --break-system-packages
  (opcional, pra mensagem personalizada) Ollama rodando localmente

USO:
  python3 jobradar.py
"""

import json
import os
import sys
import time
import urllib.request
import urllib.error

try:
    from jobspy import scrape_jobs
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.exit(
        "Faltam dependências. Rode:\n"
        "  pip install python-jobspy openpyxl --break-system-packages\n"
        "ou, se preferir instalar a partir do arquivo do projeto:\n"
        "  pip install -r requirements.txt --break-system-packages"
    )

# ---------------------------------------------------------------------------
# Cores (ANSI) - funcionam na maioria dos terminais Linux/Mac
# ---------------------------------------------------------------------------
class C:
    RESET = "\033[0m"
    BOLD = "\033[1m"
    CYAN = "\033[96m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    RED = "\033[91m"
    MAGENTA = "\033[95m"
    GRAY = "\033[90m"


BANNER = f"""{C.MAGENTA}{C.BOLD}
 ▄▄▄▄▄▄▄▄▄▄▄  ▄▄▄▄▄▄▄▄▄▄▄  ▄▄▄▄▄▄▄▄▄▄
▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌▐░░░░░░░░░░▌
 ▀▀▀▀▀█░█▀▀▀ ▐░█▀▀▀▀▀▀▀█░▌▐░█▀▀▀▀▀▀▀█░▌
      ▐░▌    ▐░▌       ▐░▌▐░▌       ▐░▌
      ▐░▌    ▐░▌       ▐░▌▐░█▄▄▄▄▄▄▄█░▌
      ▐░▌    ▐░▌       ▐░▌▐░░░░░░░░░░▌
      ▐░▌    ▐░▌       ▐░▌▐░█▀▀▀▀▀▀▀█░▌
      ▐░▌    ▐░▌       ▐░▌▐░▌       ▐░▌
 ▄▄▄▄▄█░▌    ▐░█▄▄▄▄▄▄▄█░▌▐░█▄▄▄▄▄▄▄█░▌▄
▐░░░░░░░▌    ▐░░░░░░░░░░░▌▐░░░░░░░░░░▌▐░▌
 ▀▀▀▀▀▀▀      ▀▀▀▀▀▀▀▀▀▀▀  ▀▀▀▀▀▀▀▀▀▀  ▀

 ▄▄▄▄▄▄▄▄▄▄▄  ▄▄▄▄▄▄▄▄▄▄▄  ▄▄▄▄▄▄▄▄▄▄   ▄▄▄▄▄▄▄▄▄▄▄  ▄▄▄▄▄▄▄▄▄▄▄
▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌▐░░░░░░░░░░▌ ▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌
▐░█▀▀▀▀▀▀▀█░▌▐░█▀▀▀▀▀▀▀█░▌▐░█▀▀▀▀▀▀▀█░▌▐░█▀▀▀▀▀▀▀█░▌▐░█▀▀▀▀▀▀▀█░▌
▐░▌       ▐░▌▐░▌       ▐░▌▐░▌       ▐░▌▐░▌       ▐░▌▐░▌       ▐░▌
▐░█▄▄▄▄▄▄▄█░▌▐░█▄▄▄▄▄▄▄█░▌▐░▌       ▐░▌▐░█▄▄▄▄▄▄▄█░▌▐░█▄▄▄▄▄▄▄█░▌
▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌▐░▌       ▐░▌▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌
▐░█▀▀▀▀█░█▀▀ ▐░█▀▀▀▀▀▀▀█░▌▐░▌       ▐░▌▐░█▀▀▀▀▀▀▀█░▌▐░█▀▀▀▀█░█▀▀
▐░▌     ▐░▌  ▐░▌       ▐░▌▐░▌       ▐░▌▐░▌       ▐░▌▐░▌     ▐░▌
▐░▌      ▐░▌ ▐░▌       ▐░▌▐░█▄▄▄▄▄▄▄█░▌▐░▌       ▐░▌▐░▌      ▐░▌
▐░▌       ▐░▌▐░▌       ▐░▌▐░░░░░░░░░░▌ ▐░▌       ▐░▌▐░▌       ▐░▌
 ▀         ▀  ▀         ▀  ▀▀▀▀▀▀▀▀▀▀   ▀         ▀  ▀         ▀

{C.RESET}{C.GRAY}   busca de vagas + mensagem personalizada, sem custo{C.RESET}
"""

CONFIG_FILE = "jobradar_config.json"
OUTPUT_FILE = "vagas_personalizadas.xlsx"

COLUNAS = ["Vaga", "Empresa", "Local", "Remoto", "Link Vaga", "Mensagem"]

# Se o Ollama falhar essa quantidade de vezes seguidas, desligamos a geração
# de mensagem pro resto da execução em vez de continuar tentando (cada
# tentativa tem timeout de até 120s, então insistir custaria muito tempo).
MAX_FALHAS_OLLAMA_SEGUIDAS = 3


# ---------------------------------------------------------------------------
# Helpers de interface
# ---------------------------------------------------------------------------
def perguntar(texto, padrao=""):
    """Faz uma pergunta de texto livre no terminal, usando `padrao` se o
    usuário apertar Enter sem digitar nada."""
    sufixo = f" {C.GRAY}(padrão: {padrao}){C.RESET}" if padrao != "" else ""
    resposta = input(f"{C.YELLOW}?{C.RESET} {texto}{sufixo}: ").strip()
    return resposta if resposta else padrao


def perguntar_bool3(texto, padrao=None):
    """Pergunta y/n/branco. Retorna True, False ou None."""
    mapa_padrao = {True: "y", False: "n", None: "branco = qualquer uma"}
    sufixo = f" {C.GRAY}(y/n, padrão: {mapa_padrao[padrao]}){C.RESET}"
    resposta = input(f"{C.YELLOW}?{C.RESET} {texto}{sufixo}: ").strip().lower()
    if resposta == "":
        return padrao
    if resposta in ("y", "yes", "s", "sim"):
        return True
    if resposta in ("n", "no", "nao", "não"):
        return False
    print(f"{C.RED}  Não entendi, usando o padrão.{C.RESET}")
    return padrao


def perguntar_int(texto, padrao, minimo=None):
    """Pergunta um número inteiro. Se a entrada estiver vazia, não for um
    número válido, ou ficar abaixo de `minimo` (quando informado), cai de
    volta pro padrão em vez de deixar um valor inválido seguir adiante."""
    resposta = input(f"{C.YELLOW}?{C.RESET} {texto} {C.GRAY}(padrão: {padrao}){C.RESET}: ").strip()
    if resposta == "":
        return padrao
    try:
        valor = int(resposta)
    except ValueError:
        print(f"{C.RED}  Número inválido, usando o padrão.{C.RESET}")
        return padrao
    if minimo is not None and valor < minimo:
        print(f"{C.RED}  Valor muito baixo, usando o padrão.{C.RESET}")
        return padrao
    return valor


def linha(char="─", cor=C.GRAY):
    """Imprime uma linha divisória simples, só pra organizar a saída."""
    print(f"{cor}{char * 60}{C.RESET}")


# ---------------------------------------------------------------------------
# Configuração (perguntas interativas com defaults, salva em arquivo)
# ---------------------------------------------------------------------------
def carregar_config_salva():
    """Tenta carregar a configuração da execução anterior. Se o arquivo
    não existir, não for JSON válido, ou não puder ser lido (permissão,
    disco, etc.), trata como "sem configuração salva" em vez de travar
    o script - o usuário só vai precisar configurar de novo."""
    if not os.path.exists(CONFIG_FILE):
        return None
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError, UnicodeDecodeError) as e:
        print(
            f"{C.YELLOW}Não consegui ler {CONFIG_FILE} ({e}). "
            f"Vou ignorar e pedir a configuração de novo.{C.RESET}"
        )
        return None


def montar_config():
    """Monta o dicionário de configuração da busca, reaproveitando a
    configuração salva (se o usuário topar) ou perguntando tudo de novo."""
    salva = carregar_config_salva()

    if salva:
        print(f"{C.GREEN}Encontrei uma configuração salva de execução anterior.{C.RESET}")
        usar = perguntar_bool3("Usar a mesma configuração de novo?", padrao=True)
        if usar:
            return salva
        print(f"{C.GRAY}Ok, vamos configurar de novo.{C.RESET}\n")

    linha()
    print(f"{C.BOLD}Configuração da busca{C.RESET} — deixe em branco pra usar o padrão\n")

    search_term = perguntar("Cargo/termo de busca", padrao="Analista de Suporte")
    location = perguntar("Localização", padrao="São Paulo, Brasil")
    is_remote = perguntar_bool3("Somente vagas remotas?", padrao=None)
    results_wanted = perguntar_int("Quantas vagas buscar (aprox.)", padrao=100, minimo=1)

    print()
    usar_ia = perguntar_bool3(
        "Gerar mensagem personalizada com IA local (Ollama)?", padrao=False
    )

    ollama_url = "http://localhost:11434/api/generate"
    ollama_model = "llama3.2:latest"
    perfil = ""

    if usar_ia:
        ollama_url = perguntar("URL da API do Ollama", padrao=ollama_url)
        ollama_model = perguntar("Nome do modelo Ollama", padrao=ollama_model)
        print(f"{C.GRAY}Cole um resumo do seu perfil profissional (uma linha, sem quebras):{C.RESET}")
        perfil = perguntar(
            "Perfil",
            padrao="Profissional de TI buscando novas oportunidades na área.",
        )

    config = {
        "search_term": search_term,
        "google_search_term": f"vagas {search_term} {location}",
        "location": location,
        "is_remote": is_remote,
        "results_wanted": results_wanted,
        "sites": ["linkedin", "indeed", "google"],
        "usar_ia": usar_ia,
        "ollama_url": ollama_url,
        "ollama_model": ollama_model,
        "perfil": perfil,
    }

    linha()
    salvar = perguntar_bool3("Salvar essa configuração pra próxima vez?", padrao=True)
    if salvar:
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            print(f"{C.GREEN}Configuração salva em {CONFIG_FILE}{C.RESET}")
        except OSError as e:
            print(
                f"{C.YELLOW}Não consegui salvar a configuração ({e}). "
                f"Vou seguir mesmo assim, só que sem salvar.{C.RESET}"
            )

    return config


# ---------------------------------------------------------------------------
# Validação do Ollama (só roda se usar_ia = True)
# ---------------------------------------------------------------------------
def testar_ollama(config):
    """Faz uma checagem rápida (endpoint /api/tags) pra ver se o Ollama
    está no ar antes de começar a gerar mensagem vaga por vaga. Retorna
    False (sem travar o script) se algo der errado - a busca de vagas
    continua normalmente, só sem mensagem personalizada."""
    try:
        req = urllib.request.Request(
            config["ollama_url"].replace("/api/generate", "/api/tags")
        )
        with urllib.request.urlopen(req, timeout=5) as resp:
            json.loads(resp.read().decode("utf-8"))
        print(f"{C.GREEN}✓ Ollama respondendo normalmente.{C.RESET}")
        return True
    except Exception as e:
        print(
            f"{C.RED}✗ Não consegui falar com o Ollama em {config['ollama_url']} "
            f"({e}). As mensagens vão ficar em branco.{C.RESET}"
        )
        return False


# ---------------------------------------------------------------------------
# Scraper
# ---------------------------------------------------------------------------
def rodar_scraper(config):
    """Chama o python-jobspy pra buscar vagas nos sites configurados.
    Erros de rede, bloqueio por rate-limit ou mudanças nos sites de vaga
    (fora do nosso controle) são comuns aqui, então avisamos o usuário
    com uma mensagem amigável em vez de deixar o traceback cru estourar."""
    print(
        f"\n{C.CYAN}Buscando até {config['results_wanted']} vagas para "
        f"'{config['search_term']}' em '{config['location']}'...{C.RESET}"
    )
    try:
        df = scrape_jobs(
            site_name=config["sites"],
            search_term=config["search_term"],
            google_search_term=config["google_search_term"],
            location=config["location"],
            results_wanted=config["results_wanted"],
            is_remote=config["is_remote"],
            country_indeed="Brazil",
        )
    except Exception as e:
        sys.exit(
            f"{C.RED}Falha ao buscar vagas ({e}).{C.RESET}\n"
            f"{C.GRAY}Verifique sua conexão com a internet - alguns sites também "
            f"bloqueiam buscas automatizadas temporariamente se você rodar o "
            f"script muitas vezes seguidas em pouco tempo.{C.RESET}"
        )

    if df is None or df.empty:
        print(f"{C.YELLOW}Nenhuma vaga encontrada para esses critérios.{C.RESET}")
        return df

    df = df.fillna("")
    print(f"{C.GREEN}Recebi {len(df)} vagas.{C.RESET}")
    return df


def gerar_mensagem(vaga, config):
    """Monta um prompt com o perfil do candidato e os dados da vaga, e
    pede pro modelo local (via Ollama) escrever uma mensagem curta.
    Lança RuntimeError com mensagem amigável nos casos mais comuns de
    falha (modelo inexistente, Ollama fora do ar, resposta inválida) -
    quem chama decide o que fazer com o erro."""
    titulo = vaga.get("title", "") or ""
    empresa = vaga.get("company", "") or ""
    descricao = (vaga.get("description", "") or "")[:800]

    prompt = f"""Perfil do candidato:
{config['perfil']}

Vaga: {titulo}
Empresa: {empresa}
Descrição (trecho): {descricao}

Escreva uma mensagem curta (até 500 caracteres), em português brasileiro,
para enviar a um recrutador no LinkedIn puxando assunto sobre essa vaga
específica. Tom direto, profissional, sem clichê genérico. Sem saudação
tipo "Espero que esteja bem". Não invente experiência que não está no perfil.
Retorne só o texto da mensagem, nada mais."""

    body = json.dumps({
        "model": config["ollama_model"],
        "prompt": prompt,
        "stream": False,
    }).encode("utf-8")

    req = urllib.request.Request(
        config["ollama_url"],
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        if e.code == 404:
            raise RuntimeError(
                f"Modelo '{config['ollama_model']}' não encontrado no Ollama."
            )
        raise RuntimeError(f"Ollama retornou erro HTTP {e.code}.")
    except urllib.error.URLError as e:
        raise RuntimeError(f"Não consegui conectar ao Ollama ({e.reason}).")
    except json.JSONDecodeError:
        raise RuntimeError("O Ollama respondeu algo que não é JSON válido.")

    return result.get("response", "").strip()


# ---------------------------------------------------------------------------
# Planilha (com dedupe contra execuções anteriores)
# ---------------------------------------------------------------------------
def carregar_links_existentes():
    """Lê a planilha anterior (se existir e for compatível) e retorna
    (workbook, sheet, set de links já vistos).

    Se o arquivo existir mas tiver um cabeçalho diferente (de uma versão
    antiga do script, por exemplo) ou estiver corrompido/ilegível, faz
    backup dele em vez de tentar reaproveitar - misturar formatos
    diferentes bagunça as colunas, e um arquivo corrompido travaria o
    load_workbook.
    """
    if os.path.exists(OUTPUT_FILE):
        try:
            wb_antigo = load_workbook(OUTPUT_FILE)
            ws_antigo = wb_antigo.active
            cabecalho_atual = [c.value for c in ws_antigo[1]]
        except Exception as e:
            backup_name = OUTPUT_FILE.replace(".xlsx", f"_backup_{int(time.time())}.xlsx")
            os.rename(OUTPUT_FILE, backup_name)
            print(
                f"{C.YELLOW}Não consegui abrir a planilha existente ({e}) - parece "
                f"estar corrompida ou aberta em outro programa. Guardei ela em "
                f"{backup_name} e vou começar uma nova.{C.RESET}"
            )
        else:
            if cabecalho_atual == COLUNAS:
                links = set()
                idx_link = COLUNAS.index("Link Vaga")
                for row in ws_antigo.iter_rows(min_row=2, values_only=True):
                    if len(row) > idx_link and row[idx_link]:
                        links.add(row[idx_link])
                print(f"{C.GRAY}Planilha anterior encontrada: {len(links)} vagas já registradas.{C.RESET}")
                return wb_antigo, ws_antigo, links

            # Cabeçalho não bate (versão antiga do script) - faz backup e começa do zero
            backup_name = OUTPUT_FILE.replace(".xlsx", f"_backup_{int(time.time())}.xlsx")
            os.rename(OUTPUT_FILE, backup_name)
            print(
                f"{C.YELLOW}A planilha existente tinha um formato diferente (versão antiga do script). "
                f"Guardei ela em {backup_name} e vou começar uma nova.{C.RESET}"
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "Vagas"
    ws.append(COLUNAS)
    return wb, ws, set()


def salvar_planilha(wb):
    """Salva o workbook em OUTPUT_FILE. Se o arquivo estiver aberto em
    outro programa (Excel/LibreOffice travando o arquivo) ou não puder
    ser escrito por algum outro motivo, salva num nome alternativo em
    vez de perder o trabalho já feito na execução. Retorna o caminho
    onde efetivamente salvou."""
    try:
        wb.save(OUTPUT_FILE)
        return OUTPUT_FILE
    except OSError as e:
        alternativo = OUTPUT_FILE.replace(".xlsx", f"_{int(time.time())}.xlsx")
        print(
            f"{C.RED}Não consegui salvar em {OUTPUT_FILE} ({e}). "
            f"Isso costuma acontecer quando o arquivo está aberto em outro "
            f"programa - feche-o antes da próxima execução. Salvando em "
            f"{alternativo} pra não perder o que já foi buscado.{C.RESET}"
        )
        wb.save(alternativo)
        return alternativo


def main():
    """Fluxo principal: monta a configuração, busca as vagas, gera
    mensagem (opcional) para cada uma e salva tudo na planilha, sem
    duplicar vagas já registradas em execuções anteriores."""
    print(BANNER)
    config = montar_config()

    usar_ia = config.get("usar_ia", False)
    if usar_ia:
        usar_ia = testar_ollama(config)

    linha()
    df = rodar_scraper(config)
    vagas = df.to_dict(orient="records") if df is not None and not df.empty else []

    wb, ws, links_existentes = carregar_links_existentes()

    novas, puladas = 0, 0
    total = len(vagas)
    falhas_ollama_seguidas = 0

    try:
        for i, vaga in enumerate(vagas, 1):
            titulo = vaga.get("title", "") or ""
            empresa = vaga.get("company", "") or ""
            link = vaga.get("job_url", "") or ""

            if link and link in links_existentes:
                puladas += 1
                print(f"{C.GRAY}[{i}/{total}] (já registrada, pulando) {titulo} @ {empresa}{C.RESET}")
                continue

            print(f"{C.CYAN}[{i}/{total}]{C.RESET} {titulo} @ {C.MAGENTA}{empresa}{C.RESET}")

            msg = ""
            if usar_ia:
                try:
                    msg = gerar_mensagem(vaga, config)
                    falhas_ollama_seguidas = 0
                except Exception as e:
                    msg = f"ERRO ao gerar: {e}"
                    falhas_ollama_seguidas += 1
                    if falhas_ollama_seguidas >= MAX_FALHAS_OLLAMA_SEGUIDAS:
                        print(
                            f"{C.RED}Ollama falhou {falhas_ollama_seguidas} vezes seguidas "
                            f"- desligando a geração de mensagem pro resto desta execução "
                            f"(a busca de vagas continua normalmente).{C.RESET}"
                        )
                        usar_ia = False

            ws.append([
                titulo,
                empresa,
                vaga.get("location", "") or "",
                vaga.get("is_remote", "") if vaga.get("is_remote", "") != "" else "",
                link,
                msg,
            ])
            if link:
                links_existentes.add(link)
            novas += 1
            time.sleep(0.3)
    except KeyboardInterrupt:
        print(f"\n{C.YELLOW}Interrompido pelo usuário - salvando o que já foi processado...{C.RESET}")
        salvar_planilha(wb)
        raise

    caminho_salvo = salvar_planilha(wb)

    linha()
    print(f"{C.GREEN}{C.BOLD}Pronto!{C.RESET} {novas} vagas novas adicionadas, {puladas} já existiam e foram puladas.")
    print(f"Salvo em {C.CYAN}{caminho_salvo}{C.RESET}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n{C.YELLOW}Interrompido pelo usuário.{C.RESET}")
